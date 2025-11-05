import time
import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from captcha_solver import solve_captcha_from_bytes
import io
import requests

# ==== CẤU HÌNH ====
EXCEL_FILE = r"C:\Users\Admin\OneDrive\Tài liệu\HocBong.xlsx"
SHEET_SINHVIEN = "14DH"
SHEET_MONHOC = "MonHoc"
LT_PRICE = 785000
TH_PRICE = 1000000

# ==== KHỞI TẠO TRÌNH DUYỆT ====
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
wait = WebDriverWait(driver, 30)

# ==== MỞ FILE EXCEL ====
wb = openpyxl.load_workbook(EXCEL_FILE)
sheet_sv = wb[SHEET_SINHVIEN]
sheet_monhoc = wb[SHEET_MONHOC]

# ==== ĐỌC DANH SÁCH MÔN 3TC ====
mon3tc_map = {}
for row in sheet_monhoc.iter_rows(min_row=2, values_only=True):
    ten_mon, lt, th = row
    if ten_mon:
        mon3tc_map[ten_mon.strip().lower()] = {"LT": int(lt), "TH": int(th)}

# ==== XỬ LÝ TỪNG MSSV ====
for row in sheet_sv.iter_rows(min_row=2):
    mssv_cell = row[1]
    gpa_cell = row[5]  # Cột F - GPA

    if not mssv_cell.value or gpa_cell.value:
        continue  # Bỏ qua nếu MSSV trống hoặc đã có GPA

    mssv = str(mssv_cell.value).strip()
    print(f"\n🔍 Đang xử lý MSSV: {mssv}")

    driver.get("https://sinhvien.huit.edu.vn/tra-cuu-thong-tin.html")

    try:
        # Nhập MSSV
        mssv_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Nhập mã sinh viên.']")))
        mssv_input.clear()
        mssv_input.send_keys(mssv)

        # Lấy URL ảnh CAPTCHA
        captcha_img = driver.find_element(By.ID, "newcaptcha")
        src = captcha_img.get_attribute("src")
        print(f"[DEBUG] CAPTCHA src = {src}")
        full_url = src

        # Tải ảnh về và giải mã bằng mô hình
        img_bytes = io.BytesIO(requests.get(full_url).content)
        captcha = solve_captcha_from_bytes(img_bytes)

        print(f"🤖 CAPTCHA tự động: {captcha}")

        # Nhập CAPTCHA vào ô
        captcha_input = driver.find_element(By.XPATH, "//input[@placeholder='Nhập mã']")
        captcha_input.clear()
        captcha_input.send_keys(captcha)

        # Bấm nút Tra cứu
        tra_cuu_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Tra cứu')]")))
        tra_cuu_btn.click()
        time.sleep(1)

        # 🆕 Kiểm tra lỗi CAPTCHA
        try:
            toast = driver.find_element(By.CLASS_NAME, "toast-message")
            if "Mã bảo vệ không hợp lệ" in toast.text:
                print("❌ CAPTCHA sai, bỏ qua MSSV này.")
                continue
        except:
            print("✅ CAPTCHA hợp lệ hoặc không có thông báo lỗi.")
        # Kiểm tra lỗi
        try:
            error_box = driver.find_element(By.CLASS_NAME, "alert-danger")
            print(f"❌ Trang báo lỗi: {error_box.text.strip()}")
            continue
        except:
            print("✅ Không có lỗi trả về, tiếp tục...")

        # Click XEM ĐIỂM
        xem_diem_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Xem điểm')]")))
        xem_diem_btn.click()

        # Chuyển sang tab mới
        driver.switch_to.window(driver.window_handles[-1])
        time.sleep(2)

        # Phân tích bảng điểm
        soup = BeautifulSoup(driver.page_source, "html.parser")
        table = soup.find("table", {"id": "xemDiem"})
        if not table:
            print("❌ Không tìm thấy bảng điểm.")
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            continue

        # Tìm học kỳ HK2 (2024 - 2025)
        hk2_td = soup.find("td", string=lambda s: s and "HK2 (2024" in s)
        if not hk2_td:
            print("⚠️ Không có học kỳ HK2. Ghi điểm = 0.")
            row[4].value = 0  # TC
            row[5].value = 0.0  # GPA
            row[7].value = 0  # Học phí
            wb.save(EXCEL_FILE)
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            continue

        # Phân tích từng môn
        start_tr = hk2_td.find_parent("tr")
        rows = []
        for tr in start_tr.find_next_siblings("tr"):
            if tr.text.strip().startswith("Tính theo"):
                break
            if not tr.find_all("td", recursive=False):
                continue
            rows.append(tr)

        tong_tc = tong_diem = tong_hocphi = 0
        for tr in rows:
            cols = tr.find_all("td")
            if len(cols) < 8:
                continue

            ten_mon = cols[2].text.strip()
            ten_mon_lower = ten_mon.lower()

            # Bỏ qua thể chất và quốc phòng
            if "thể chất" in ten_mon_lower or "quốc phòng" in ten_mon_lower:
                print(f"⚠️ Bỏ qua môn: {ten_mon}")
                continue

            try:
                so_tc = int(cols[3].text.strip())
            except:
                print(f"⚠️ Không đọc được số TC: {ten_mon}")
                continue

            diem = 0.0
            for td in cols:
                if td.get("title") == "DiemTongKet":
                    diem_raw = td.text.strip()
                    try:
                        diem = float(diem_raw.replace(",", "."))
                    except:
                        print(f"⚠️ Lỗi điểm: {diem_raw}")
                    break

            # Xác định LT - TH
            if so_tc == 1:
                lt, th = 0, 1
            elif so_tc == 2:
                lt, th = 2, 0
            elif so_tc == 3 and ten_mon_lower in mon3tc_map:
                lt = mon3tc_map[ten_mon_lower]["LT"]
                th = mon3tc_map[ten_mon_lower]["TH"]
            else:
                lt, th = so_tc, 0

            hoc_phi = lt * LT_PRICE + th * TH_PRICE
            tong_tc += so_tc
            tong_diem += diem * so_tc
            tong_hocphi += hoc_phi

            print(f"✅ {ten_mon} | TC: {so_tc} | Điểm: {diem} | HP: {hoc_phi}")

        gpa = round(tong_diem / tong_tc, 2) if tong_tc else 0.0
        row[4].value = tong_tc
        row[5].value = gpa
        row[7].value = tong_hocphi
        wb.save(EXCEL_FILE)

        print(f"✅ GPA: {gpa}, TC: {tong_tc}, HP: {tong_hocphi}")

        driver.close()
        driver.switch_to.window(driver.window_handles[0])

    except Exception as e:
        print(f"❌ Lỗi MSSV {mssv}: {e.__class__.__name__}: {e}")
        if len(driver.window_handles) > 1:
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
        continue

# ==== KẾT THÚC ====
driver.quit()
print("\nĐã cập nhật xong Excel:", EXCEL_FILE)
